from customtkinter import *
import random
import openpyxl
import datetime



#region Janela
janela = CTk()
janela.title('Gerador de Códigos')
janela.geometry("835x650")
janela.resizable(width=False, height=False)
#endregion Janela


#region Funções

def ValoresParaGerar(Serie, QuantidadeDeDigitos, QuantidadeDeCodigos):
    caracteres = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789'
    codigos_gerados = set()
    quantidade_gerada = 0

    resultados.configure(state="normal")
    resultados.delete(1.0, END)

    while quantidade_gerada < QuantidadeDeCodigos:
        cod = [str(Serie)]
        for _ in range(QuantidadeDeDigitos):
            cod.append(random.choice(caracteres))
        codigo_gerado = ''.join(cod)
        if codigo_gerado not in codigos_gerados:
            resultados.insert(END, f"{codigo_gerado}\n")
            codigos_gerados.add(codigo_gerado)
            quantidade_gerada += 1

    resultados.configure(state="disabled")

def gerar_codigos():
    erro_label.configure(text="")

    try:
        serie = serie_entry.get()
        quantdigitos = int(quantdigitos_entry.get())
        quantidadecods = int(quantidadecods_entry.get())
        if quantdigitos <= 0 or quantidadecods <= 0:
            erro_label.configure(text="Quantidade de dígitos e de códigos devem ser maiores que zero.")
            return

        ValoresParaGerar(serie, quantdigitos, quantidadecods)
    except ValueError:
        erro_label.configure(text="Por favor, insira valores numéricos válidos.")


def limpar_texto():
    resultados.configure(state="normal")
    resultados.delete(1.0, END)
    resultados.configure(state="disabled")


def copiar_codigos():
    texto = resultados.get(1.0, END).strip().replace('\n', '\n')
    janela.clipboard_clear()
    janela.clipboard_append(texto)


def formatar():
    # Obter o conteúdo atual do TextBox
    resultados.configure(state="normal")
    conteudo = resultados.get(1.0, END).strip().split('\n')

    # Limpar o TextBox
    resultados.delete(1.0, END)

    # Formatar cada linha e inserir de volta no TextBox
    for linha in conteudo:
        if not linha.startswith('codigos.Add("') and not linha.endswith('");'):
            linha_formatada = f'codigos.Add("{linha}");'
            resultados.insert(END, f"{linha_formatada}\n")
        else:
            resultados.insert(END, f"{linha}\n")
    resultados.configure(state="disabled")


def desformatar():
    resultados.configure(state="normal")
    conteudo = resultados.get(1.0, END).strip().split('\n')
    resultados.delete(1.0, END)
    for linha in conteudo:
        if linha.startswith('codigos.Add("') and linha.endswith('");'):
            linha_desformatada = linha[len('codigos.Add("'):-len('");')]
            resultados.insert(END, f"{linha_desformatada}\n")
        else:
            resultados.insert(END, f"{linha}\n")

    resultados.configure(state="disabled")


def GerarPlanilha():
    book = openpyxl.Workbook()
    book.create_sheet(f'Códigos {datetime.date.today()}')
    pagCodes = book[f'Códigos {datetime.date.today()}']

    conteudo = resultados.get(1.0, END).strip().split('\n')

    linha=2

    bold = openpyxl.styles.Font(bold=True, name='Arial', size=15)
    alinharCentro= openpyxl.styles.Alignment(horizontal='center', vertical='center')
    borda = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
    fundo = openpyxl.styles.PatternFill(start_color="E87516", end_color="FFFF99", fill_type="solid")

    pagCodes.cell(row=1, column=1, value='Código Desformatado:').font=bold
    pagCodes.cell(row=1, column=2, value='Código Formatado:').font=bold
    pagCodes.cell(row=1, column=1).alignment = alinharCentro
    pagCodes.cell(row=1, column=2).alignment = alinharCentro
    pagCodes.cell(row=1, column=1).border = borda
    pagCodes.cell(row=1, column=2).border = borda
    pagCodes.cell(row=1, column=1).fill = fundo
    pagCodes.cell(row=1, column=2).fill = fundo

    for codigo in conteudo:
        pagCodes.cell(row=linha, column=1, value=codigo)
        pagCodes.cell(row=linha, column=2, value=f'codigos.Add("{codigo}");')
        pagCodes.cell(row=linha, column=1).font = bold
        pagCodes.cell(row=linha, column=2).font = bold
        pagCodes.cell(row=linha, column=1).border = borda
        pagCodes.cell(row=linha, column=2).border = borda
        pagCodes.cell(row=linha, column=1).alignment = alinharCentro
        pagCodes.cell(row=linha, column=2).alignment = alinharCentro
        linha+=1

    pagCodes.column_dimensions['A'].width = 45
    pagCodes.column_dimensions['B'].width = 45

    pagCodes.row_dimensions[1].height=40

    book.save(f'Planilha {datetime.date.today()}.xlsx')


def exportar():
    exportar = CTkToplevel(janela)

    exportar.title('Exportar Códigos')
    exportar.geometry("400x400")
    exportar.resizable(width=False, height=False)

    exportar.transient(janela)
    exportar.grab_set()

    txt_copiar = CTkLabel(exportar, text="Copiar os códigos\nseparados por linha", font=("Arial", 30, "bold"))
    txt_copiar.place(x=60, y=20)
    botao_copiar = CTkButton(exportar, text="COPIAR CÓDIGOS", font=("Arial", 30, "bold"), command=copiar_codigos,
                             width=280, height=60, fg_color="white", text_color="Black", hover_color="gray")
    botao_copiar.place(x=60, y=100)

    txt_excel = CTkLabel(exportar, text="Gerar planilha com\nos códigos", font=("Arial", 30, "bold"))
    txt_excel.place(x=60, y=220)
    botao_gerar_excel = CTkButton(exportar, command=GerarPlanilha, text="GERAR PLANILHA", font=("Arial", 30, "bold"), width=280, height=60,
                                  fg_color="black", text_color="white", hover_color="gray")
    botao_gerar_excel.place(x=60, y=300)

    exportar.mainloop()

#endregion Funções


#region Painel
txt_entry = CTkLabel(janela, text="Digite a série para os códigos", font=("Arial", 17)).place(x=10, y=15)

txt_quantdigitos = CTkLabel(janela, text="Quantidade de dígitos por códigos", font=("Arial", 17)).place(x=285, y=15)

txt_quantidadecods_entry = CTkLabel(janela, text="Quantidade de códigos", font=("Arial", 17)).place(x=560, y=15)

serie_entry = CTkEntry(janela, font=("Arial", 15), width=265, height=50)
serie_entry.place(x=10, y=50)

quantdigitos_entry = CTkEntry(janela, font=("Arial", 15), width=265, height=50)
quantdigitos_entry.place(x=285, y=50)

quantidadecods_entry = CTkEntry(janela, font=("Arial", 15), width=265, height=50)
quantidadecods_entry.place(x=560, y=50)

botao_gerar = CTkButton(janela, text="GERAR", font=("Arial", 30, "bold"), command=gerar_codigos, width=200, height=60)
botao_gerar.place(x=107.5, y=110),

botao_exportar = CTkButton(janela, text="EXPORTAR", font=("Arial", 30, "bold"), command=exportar, width=200, height=60,
                           fg_color="green", hover_color="dark green")
botao_exportar.place(x=317.5, y=110)

botao_limpar = CTkButton(janela, text="LIMPAR", font=("Arial", 30, "bold"), command=limpar_texto, width=200, height=60,
                         fg_color="red", hover_color="dark red")
botao_limpar.place(x=527.5, y=110)

botao_formatar = CTkButton(janela, command=formatar, text='FORMATAR', fg_color='Black', hover_color='gray',
                           font=("Arial", 30, 'bold'), width=200, height=60)
botao_formatar.place(x=613.75, y=330)

botao_desformatar = CTkButton(janela, command=desformatar, text='DESFORMATAR', fg_color='Black', hover_color='gray',
                              font=("Arial", 25, 'bold'), width=200, height=60)
botao_desformatar.place(x=613.75, y=410)

resultados = CTkTextbox(janela, font=("Arial", 30), width=571.25, height=400, state="disabled")
resultados.place(x=21.25, y=200)

erro_label = CTkLabel(janela, text="", font=("Arial", 20), text_color="red")
erro_label.place(x=67.5, y=610)

#endregion Painel


janela.mainloop()