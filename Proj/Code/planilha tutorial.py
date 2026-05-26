import openpyxl
import datetime

book=openpyxl.Workbook()
print(book.sheetnames)

book.create_sheet(f'Códigos {datetime.date.today()}')
pagCodes=book[f'Códigos {datetime.date.today()}']
pagCodes.append([input('caracter: ')])
pagCodes.append([input('caracter: ')])
pagCodes.append([input('caracter: ')])
book.save('planilha.xlsx')